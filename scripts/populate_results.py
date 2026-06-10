#!/usr/bin/env nix-shell
#!nix-shell -p python312Packages.openpyxl -i python3
"""
Final xlsx population with matching layout:
  - Row 2: "With Heuristics" (G-AO), "Without Heuristics" (V-AT)
  - Row 3: "Time (µs)" sub-header over timing cols, "Memory (B)" over memory cols
  - Row 4: per-column labels with units
  - Rows 5-24: data

Layout:
  With Heuristics:   G-H (cells filled) | I-N (graph stats) | O-U (time µs) | AK-AO (memory B)
  Without Heuristics: V-W (cells filled) | X-AC (graph stats) | AD-AJ (time µs) | AP-AT (memory B)
"""
import csv
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CSV = 'results/experiment_results.csv'
XLSX = 'results/experiment_results.xlsx'

ROW_MAP = {
    '9x9':  {'very_easy':5,'easy':6,'medium':7,'hard':8,'very_hard':9,'ambiguous':10},
    '16x16': {'very_easy':12,'easy':13,'medium':14,'hard':15,'very_hard':16,'ambiguous':17},
    '25x25': {'very_easy':19,'easy':20,'medium':21,'hard':22,'very_hard':23,'ambiguous':24},
}

WITH_COLS = [
    ('G','cells_filled'),('H','pct_solved'),('I','initial_vertices'),('J','initial_edges'),
    ('K','removed_vertices'),('L','edges_removed'),('M','pct_removed_verts'),('N','pct_removed_edges'),
    ('O','mask_us'),('P','heuristic_us'),('Q','perm_us'),('R','graph_us'),
    ('S','prune_us'),('T','extract_us'),('U','total_us'),
]
WITHOUT_COLS = [
    ('V','cells_filled'),('W','pct_solved'),('X','initial_vertices'),('Y','initial_edges'),
    ('Z','removed_vertices'),('AA','edges_removed'),('AB','pct_removed_verts'),('AC','pct_removed_edges'),
    ('AD','mask_us'),('AE','heuristic_us'),('AF','perm_us'),('AG','graph_us'),
    ('AH','prune_us'),('AI','extract_us'),('AJ','total_us'),
]
MEM_H_ON  = [('AK','mask_mem'),('AL','heuristic_mem'),('AM','perm_mem'),('AN','graph_mem'),('AO','prune_mem')]
MEM_H_OFF = [('AP','mask_mem'),('AQ','heuristic_mem'),('AR','perm_mem'),('AS','graph_mem'),('AT','prune_mem')]

TIME_NS = ['mask_ns','heuristic_ns','perm_ns','graph_ns','prune_ns','extract_ns','total_ns']

HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
OOM_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
OOM_FONT = Font(color="B71C1C", bold=True, size=9)
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def fmt_mem(v):
    """Auto-scale byte values: B for <1KB, KB for <1MB, MB otherwise."""
    if v is None or v <= 0:
        return '0 B'
    if v < 1024:
        return f'{v:.0f} B'
    if v < 1024 ** 2:
        return f'{v / 1024:.1f} KB'
    return f'{v / 1024 ** 2:.1f} MB'

def avg(grp, field):
    vals = [float(r[field]) for r in grp if r.get(field) and float(r.get(field,0)) > 0]
    return sum(vals)/len(vals) if vals else 0.0

def rnd(v):
    if abs(v) >= 100: return round(v,1)
    if abs(v) >= 10: return round(v,2)
    if abs(v) >= 1: return round(v,3)
    return round(v,4) if v>0 else 0.0

# ======== COMPUTE DATA ========

rows = []
with open(CSV) as f:
    for r in csv.DictReader(f):
        for tf in TIME_NS:
            if tf in r and r[tf]:
                r[tf.replace('_ns','_us')] = str(float(r[tf])/1000.0)
        rows.append(r)

groups = defaultdict(list)
for r in rows:
    groups[(r['size'], r['category'], r['heuristic_on'])].append(r)

data = {}
for (size, cat, h_on), grp in groups.items():
    if size not in ROW_MAP or cat not in ROW_MAP[size]:
        continue
    rn = ROW_MAP[size][cat]
    n = len(grp)
    data[(rn,'E')] = n
    data[(rn,'F')] = round(avg(grp,'clues'), 1)

    cols = WITH_COLS if h_on=='true' else WITHOUT_COLS
    for col, field in cols:
        if field == 'cells_filled':
            val = avg(grp,'cells_filled') if h_on=='true' else 0
        elif field == 'pct_solved':
            total = sum(int(r.get('cells_filled','0')) for r in grp)
            s = 81 if size=='9x9' else (256 if size=='16x16' else 625)
            empty = s*n - sum(float(r['clues']) for r in grp if float(r.get('clues',0))>0)
            val = total/empty*100 if empty>0 else 0
        elif field == 'edges_removed':
            val = sum(int(r['initial_edges']) for r in grp) - sum(int(r['pruned_edges']) for r in grp)
        elif field == 'pct_removed_verts':
            rem = sum(int(r['removed_vertices']) for r in grp)
            init = sum(int(r['initial_vertices']) for r in grp)
            val = rem/init*100 if init>0 else 0
        elif field == 'pct_removed_edges':
            init = sum(int(r['initial_edges']) for r in grp)
            pruned = sum(int(r['pruned_edges']) for r in grp)
            val = (init-pruned)/init*100 if init>0 else 0
        else:
            val = avg(grp, field)
        data[(rn,col)] = rnd(val)

    mem_cols = MEM_H_ON if h_on=='true' else MEM_H_OFF
    for col, field in mem_cols:
        data[(rn,col)] = round(avg(grp,field), 1)

# ======== OOM DETECTION ========
# A row is OOM-only if the h=off group has zero successful solves
# (everything OOM-killed or panicked). Those cells get the 'OOM' marker.
oom_only_rows = set()
for (size, cat, h), grp in groups.items():
    if h != 'false':
        continue
    if size not in ROW_MAP or cat not in ROW_MAP[size]:
        continue
    successful = [r for r in grp if r.get('classification') not in ('OOM', 'Panic')]
    if not successful and grp:
        oom_only_rows.add(ROW_MAP[size][cat])

# Apply OOM marker to all h=off cells in OOM-only rows (timing + memory)
oom_col_letters = [c for c, _ in WITHOUT_COLS] + [c for c, _ in MEM_H_OFF]
for rn in oom_only_rows:
    for col in oom_col_letters:
        data[(rn, col)] = 'OOM'

# Format memory cells with auto-scaled units (B / KB / MB)
mem_col_letters = [c for c, _ in MEM_H_ON] + [c for c, _ in MEM_H_OFF]
for (rn, col), val in list(data.items()):
    if col in mem_col_letters and val != 'OOM':
        data[(rn, col)] = fmt_mem(val)

# ======== WRITE XLSX ========

wb = load_workbook(XLSX)
ws = wb.active

# Unmerge all existing merged cells in rows 2-4
for merged_range in list(ws.merged_cells.ranges):
    min_row = merged_range.min_row
    if min_row <= 4:
        ws.unmerge_cells(str(merged_range))

# --- ROW 2: Section headers ---
ws['G2'] = 'With Heuristics'
ws['V2'] = 'Without Heuristics'
ws['AK2'] = 'With Heuristics'
ws['AP2'] = 'Without Heuristics'

# Merge cells for section headers
ws.merge_cells('G2:U2')    # stats + timing
ws.merge_cells('V2:AJ2')   # stats + timing
ws.merge_cells('AK2:AO2')  # memory
ws.merge_cells('AP2:AT2')  # memory

for cell in [ws['G2'], ws['V2'], ws['AK2'], ws['AP2']]:
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = HEADER_FILL
    cell.border = BORDER

# --- ROW 3: Sub-section headers ---
ws['O3'] = 'Time (Average) for Each Phase'
ws['AD3'] = 'Time (Average) for Each Phase'
ws['AK3'] = 'Memory (Average) for Each Phase'
ws['AP3'] = 'Memory (Average) for Each Phase'

ws.merge_cells('O3:U3')
ws.merge_cells('AD3:AJ3')
ws.merge_cells('AK3:AO3')
ws.merge_cells('AP3:AT3')

for cell in [ws['O3'], ws['AD3'], ws['AK3'], ws['AP3']]:
    cell.font = Font(bold=True, size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

# Also add sub-headers for graph stats sections
ws['G3'] = 'Graph Statistics'
ws['V3'] = 'Graph Statistics'
ws.merge_cells('G3:N3')
ws.merge_cells('V3:AC3')
for cell in [ws['G3'], ws['V3']]:
    cell.font = Font(bold=True, size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

# --- ROW 4: Column headers with units ---
row4_labels = {
    'E4': 'Puzzle count', 'F4': 'Avg clues given',
    'G4': 'Cells filled', 'H4': 'Cells filled (%)',
    'I4': 'Permutations', 'J4': 'Edges added',
    'K4': 'Perms removed', 'L4': 'Edges removed',
    'M4': 'Perms removed (%)', 'N4': 'Edges removed (%)',
    'O4': 'Mask', 'P4': 'Heuristic', 'Q4': 'Permutation',
    'R4': 'Graph', 'S4': 'Pruning', 'T4': 'Extraction', 'U4': 'Total',
    'V4': 'Cells filled', 'W4': 'Cells filled (%)',
    'X4': 'Permutations', 'Y4': 'Edges added',
    'Z4': 'Perms removed', 'AA4': 'Edges removed',
    'AB4': 'Perms removed (%)', 'AC4': 'Edges removed (%)',
    'AD4': 'Mask', 'AE4': 'Heuristic', 'AF4': 'Permutation',
    'AG4': 'Graph', 'AH4': 'Pruning', 'AI4': 'Extraction', 'AJ4': 'Total',
    'AK4': 'Mask (B/KB/MB)', 'AL4': 'Heuristic (B/KB/MB)', 'AM4': 'Permutation (B/KB/MB)',
    'AN4': 'Graph (B/KB/MB)', 'AO4': 'Prune (B/KB/MB)',
    'AP4': 'Mask (B/KB/MB)', 'AQ4': 'Heuristic (B/KB/MB)', 'AR4': 'Permutation (B/KB/MB)',
    'AS4': 'Graph (B/KB/MB)', 'AT4': 'Prune (B/KB/MB)',
}
for ref, label in row4_labels.items():
    ws[ref] = label
    ws[ref].font = Font(bold=True, size=9)
    ws[ref].alignment = Alignment(horizontal='center', wrap_text=True)
    ws[ref].border = BORDER

# --- DATA ROWS 5-24 ---
for (rn, col_letter), val in data.items():
    cell = ws[f'{col_letter}{rn}']
    cell.value = val
    cell.alignment = Alignment(horizontal='right')
    cell.border = BORDER
    if val == 'OOM':
        cell.fill = OOM_FILL
        cell.font = OOM_FONT
        cell.alignment = Alignment(horizontal='center')

# Column widths
ws.column_dimensions['D'].width = 20
for col_letter in 'EFGHIJKLMNOPQRSTUVWXYZ':
    ws.column_dimensions[col_letter].width = 14
for col_letter in ['AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT']:
    ws.column_dimensions[col_letter].width = 14

wb.save(XLSX)
print(f"Wrote {len(data)} data cells to {XLSX}")
